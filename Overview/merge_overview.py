#update variables
#data_A
#data_C
#'A','B','C'
#last_non_nan_column_A
#last_non_nan_column_C
#overview_sheet_a
#overview_sheet_c
#overview A
#overview c
#C:\Users\nor_t\OneDrive\Desktop\test\Aggregated_Data.xlsx
##################
import pandas as pd
from openpyxl import Workbook, load_workbook
import os

# Ask the user for the folder containing the Excel files
folder_path = input("Please provide the folder directory containing the Excel files: ")

# Function to extract data from a specific Excel workbook
def extract_overviewdata_from_workbook(file_path):
    # Read all sheets into a dictionary of DataFrames
    df_dict = pd.read_excel(file_path, sheet_name=None)
    data_A = None
    data_C = None
    
    for sheet_name, df in df_dict.items():
        # Find the start of each table
        table_starts = df[df.iloc[:, 0].isin(['A','B','C'])].index.tolist()
        table_starts.append(len(df))  # Add the end of the DataFrame as the end of the last table
        
        # Extract and filter table A and C
        for i in range(len(table_starts) - 1):
            start_row = table_starts[i] #start of the table
            end_row = table_starts[i + 1] #start of the next table, next index of table_starts
            table = df.iloc[start_row:end_row].reset_index(drop=True)

            if i == 0 : #First overview table
                # Filter the table to include only rows where the first column matches 'Pass', 'Fails', 'Status(%)', 'Version'
                filtered_table = table[table.iloc[:, 0].isin(['Pass', 'Fails', 'Status(%)', 'Version'])]
                if not filtered_table.empty:
                    data_A = filtered_table.apply(lambda col: col[col.last_valid_index()], axis=1)
            if i == 2 : #Third overview table
                filtered_table = table[table.iloc[:, 0].isin(['Pass', 'Fails', 'Status(%)', 'Version'])]
                if not filtered_table.empty:
                    data_C = filtered_table.apply(lambda col: col[col.last_valid_index()], axis=1)
        
    return data_A, data_C

# Create a new workbook and add "Overview A" and "Overview C" sheets
new_workbook = Workbook()

# Add "Overview A" sheet
overview_sheet_a = new_workbook.active
overview_sheet_a.title = "Overview A"
header = ["AWS Foundational Security Practice"]
subheadings = ["Title", "Pass", "Fails", "Status(%)", "Version"]
overview_sheet_a.append(header)
overview_sheet_a.append(subheadings)

# Add the "Overview C" sheet
overview_sheet_c = new_workbook.create_sheet(title="Overview C")
header2 = ["Best Practice"]
overview_sheet_c.append(header2)
overview_sheet_c.append(subheadings)


# Iterate through all Excel files in the specified folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        data_A, data_C = extract_overviewdata_from_workbook(file_path)
        
        workbook_name = os.path.basename(file_path)
        if data_A is not None:
            overview_sheet_a.append([workbook_name, *data_A.tolist()])
        if data_C is not None:
            overview_sheet_c.append([workbook_name, *data_C.tolist()])
        if data_A is None and data_C is None:
            print(f"File {filename} does not contain relevant data.") 
            
#Save the new workbook
save_path = os.path.join(folder_path, "Aggregated_Data.xlsx")
new_workbook.save(save_path)
print("Data aggregated and saved to Aggregated_Data.xlsx")